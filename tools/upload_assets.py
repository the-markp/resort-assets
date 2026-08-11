#!/usr/bin/env python3
"""
upload_assets.py — G-Tracker bulk asset upload script

Reads a filled-in Excel template and uploads each row as a new asset
via the G-Tracker API.

Usage:
  python upload_assets.py --file gtracker_asset_upload_template.xlsx
  python upload_assets.py --file assets.xlsx --url https://g.endonalaw.com
  python upload_assets.py --file assets.xlsx --dry-run
  python upload_assets.py --file assets.xlsx --skip

Requirements:
  pip install openpyxl requests
"""

import argparse
import getpass
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

# ── Constants ─────────────────────────────────────────────────────────────────
VALID_CATEGORIES = {
    "rooms_facilities", "furniture_equipment", "vehicles_transport",
    "it_electronics", "maintenance_tools", "inventory_consumables",
}
VALID_STATUSES = {"available", "in_use", "maintenance", "retired", "lost"}
VALID_DEPR     = {"straight_line", "declining_balance", "custom_rate", "none"}

# Maps Excel column header → API field name
# (handles case/space variations in case users edit the header)
COLUMN_MAP = {
    "asset_number":           "asset_number",
    "asset number":           "asset_number",
    "name":                   "name",
    "category":               "category",
    "status":                 "status",
    "location":               "location",
    "serial_number":          "serial_number",
    "serial number":          "serial_number",
    "accountable_department": "accountable_department",
    "accountable department": "accountable_department",
    "accountable_person":     "accountable_person",
    "accountable person":     "accountable_person",
    "purchase_date":          "purchase_date",
    "purchase date":          "purchase_date",
    "purchase_value":         "purchase_value",
    "purchase value":         "purchase_value",
    "service_life_years":     "service_life_years",
    "service life years":     "service_life_years",
    "service life (years)":   "service_life_years",
    "depreciation_method":    "depreciation_method",
    "depreciation method":    "depreciation_method",
    "depreciation_rate":      "depreciation_rate",
    "depreciation rate":      "depreciation_rate",
    "repair_cost":            "repair_cost",
    "repair cost":            "repair_cost",
    "notes":                  "notes",
}


# ── Colours for terminal output ───────────────────────────────────────────────
def green(s):  return f"\033[92m{s}\033[0m"
def red(s):    return f"\033[91m{s}\033[0m"
def yellow(s): return f"\033[93m{s}\033[0m"
def bold(s):   return f"\033[1m{s}\033[0m"
def dim(s):    return f"\033[2m{s}\033[0m"


# ── Auth ──────────────────────────────────────────────────────────────────────
def login(base_url: str) -> str:
    """Prompt for credentials and return a JWT token."""
    print(bold("\nG-Tracker Login"))
    username = input("  Username: ").strip()
    password = getpass.getpass("  Password: ")

    try:
        r = requests.post(
            f"{base_url}/api/auth/login",
            data={"username": username, "password": password},
            timeout=10,
        )
        r.raise_for_status()
        token = r.json()["access_token"]
        user  = r.json()["user"]
        print(green(f"  ✓ Logged in as {user['full_name'] or user['username']} ({user['role']})"))
        return token
    except requests.exceptions.ConnectionError:
        print(red(f"\n  ✗ Cannot connect to {base_url}"))
        print(dim("    Check the --url argument and ensure the server is running."))
        sys.exit(1)
    except requests.HTTPError as e:
        print(red(f"\n  ✗ Login failed: {e.response.json().get('detail', str(e))}"))
        sys.exit(1)


# ── Excel reading ─────────────────────────────────────────────────────────────
def read_sheet(filepath: str, sheet_name: str) -> list[dict]:
    """Read the Excel template and return a list of raw row dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(red(f"\n✗ Sheet '{sheet_name}' not found in {filepath}"))
        print(f"  Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    ws = wb[sheet_name]

    # Read headers from row 2 (row 1 is the title)
    headers = []
    for cell in ws[2]:
        if cell.value:
            # Strip asterisks and whitespace from header labels
            raw = str(cell.value).lower().replace("*", "").strip()
            headers.append(raw)
        else:
            headers.append(None)

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip the example row (row 3) if it still contains the example value
        # Row 3 is the example row — skip if name cell (col B, index 1) still has example value
        if row_idx == 3 and str(row[1] or "").strip() == "Deluxe Room 201":
            continue
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_dict = {"_row": row_idx}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = value
        rows.append(row_dict)

    return rows


# ── Validation ────────────────────────────────────────────────────────────────
def validate_row(raw: dict) -> tuple[dict, list[str]]:
    """
    Validate a raw row dict. Returns (payload, errors).
    payload is ready to POST if errors is empty.
    """
    errors = []
    row_num = raw.get("_row", "?")

    def get(key):
        # Look up by mapped field name or direct key
        for raw_key, mapped in COLUMN_MAP.items():
            if mapped == key and raw_key in raw:
                return raw[raw_key]
        return raw.get(key)

    # ── Optional: asset_number (unique, auto-generated if blank) ────────────
    asset_number = opt_str("asset_number") if False else None  # resolved below after opt_str defined

    # ── Required fields ───────────────────────────────────────────────────────
    name = str(get("name") or "").strip()
    if not name:
        errors.append("'name' is required")

    category = str(get("category") or "").strip().lower()
    if not category:
        errors.append("'category' is required")
    elif category not in VALID_CATEGORIES:
        errors.append(f"'category' must be one of: {', '.join(sorted(VALID_CATEGORIES))}")

    status = str(get("status") or "available").strip().lower()
    if status not in VALID_STATUSES:
        errors.append(f"'status' must be one of: {', '.join(sorted(VALID_STATUSES))}")

    # ── Optional fields ───────────────────────────────────────────────────────
    purchase_date = None
    raw_date = get("purchase_date")
    if raw_date:
        if isinstance(raw_date, datetime):
            purchase_date = raw_date.strftime("%Y-%m-%d")
        else:
            ds = str(raw_date).strip()
            try:
                datetime.strptime(ds, "%Y-%m-%d")
                purchase_date = ds
            except ValueError:
                errors.append(f"'purchase_date' must be YYYY-MM-DD, got: {ds!r}")

    purchase_value = None
    raw_pv = get("purchase_value")
    if raw_pv is not None and str(raw_pv).strip() != "":
        try:
            purchase_value = str(float(str(raw_pv).replace(",", "").replace("₱", "").strip()))
        except ValueError:
            errors.append(f"'purchase_value' must be a number, got: {raw_pv!r}")

    service_life = None
    raw_sl = get("service_life_years")
    if raw_sl is not None and str(raw_sl).strip() != "":
        try:
            service_life = int(float(str(raw_sl)))
            if service_life <= 0:
                raise ValueError
        except ValueError:
            errors.append(f"'service_life_years' must be a positive integer, got: {raw_sl!r}")

    depr_method = None
    raw_dm = get("depreciation_method")
    if raw_dm is not None and str(raw_dm).strip() != "":
        depr_method = str(raw_dm).strip().lower()
        if depr_method not in VALID_DEPR:
            errors.append(f"'depreciation_method' must be one of: {', '.join(sorted(VALID_DEPR))}")

    depr_rate = None
    raw_dr = get("depreciation_rate")
    if raw_dr is not None and str(raw_dr).strip() != "":
        try:
            depr_rate = float(str(raw_dr).replace("%", "").strip())
            if not (0 < depr_rate <= 100):
                raise ValueError
        except ValueError:
            errors.append(f"'depreciation_rate' must be a number 0–100, got: {raw_dr!r}")

    repair_cost = None
    raw_rc = get("repair_cost")
    if raw_rc is not None and str(raw_rc).strip() != "":
        try:
            repair_cost = float(str(raw_rc).replace(",", "").replace("₱", "").strip())
            if repair_cost < 0:
                raise ValueError
        except ValueError:
            errors.append(f"'repair_cost' must be a non-negative number, got: {raw_rc!r}")

    def opt_str(key):
        v = get(key)
        return str(v).strip() if v is not None and str(v).strip() != "" else None

    # Resolve asset_number now that opt_str is available
    asset_number = opt_str("asset_number")
    # Validate format if supplied (no spaces, reasonable length)
    if asset_number:
        if len(asset_number) > 50:
            errors.append("'asset_number' must be 50 characters or fewer")
        if " " in asset_number:
            errors.append("'asset_number' must not contain spaces")

    payload = {
        "asset_number":           asset_number,
        "name":                   name,
        "category":               category,
        "status":                 status,
        "location":               opt_str("location"),
        "serial_number":          opt_str("serial_number"),
        "accountable_department": opt_str("accountable_department"),
        "accountable_person":     opt_str("accountable_person"),
        "purchase_date":          purchase_date,
        "purchase_value":         purchase_value,
        "service_life_years":     service_life,
        "depreciation_method":    depr_method,
        "depreciation_rate":      depr_rate,
        "repair_cost":            repair_cost,
        "notes":                  opt_str("notes"),
    }
    # Strip None values — API treats missing keys as "no change"
    payload = {k: v for k, v in payload.items() if v is not None}

    return payload, errors


# ── Upload ────────────────────────────────────────────────────────────────────
def upload_asset(base_url: str, token: str, payload: dict) -> dict:
    r = requests.post(
        f"{base_url}/api/assets/",
        json=payload,
        headers={"Authorization": f"Bearer {token}"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Bulk-upload assets to G-Tracker from an Excel file."
    )
    parser.add_argument("--file",    required=True, help="Path to the Excel file")
    parser.add_argument("--url",     default="http://localhost:8000",
                        help="G-Tracker base URL (default: http://localhost:8000)")
    parser.add_argument("--sheet",   default="Assets",
                        help="Sheet name to read from (default: Assets)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate only — do not upload")
    parser.add_argument("--skip",    action="store_true",
                        help="Skip invalid rows and continue uploading valid ones")
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print(red(f"\n✗ File not found: {filepath}"))
        sys.exit(1)

    print(bold(f"\nG-Tracker Bulk Asset Upload"))
    print(dim(f"  File  : {filepath}"))
    print(dim(f"  Sheet : {args.sheet}"))
    print(dim(f"  URL   : {args.url}"))
    if args.dry_run:
        print(yellow("  Mode  : DRY RUN (no data will be uploaded)"))

    # ── Read Excel ────────────────────────────────────────────────────────────
    print(f"\nReading {filepath.name}…")
    rows = read_sheet(str(filepath), args.sheet)
    print(f"  Found {len(rows)} data row(s) (excluding example row and empty rows)")

    if not rows:
        print(yellow("\nNo data rows found. Fill in the 'Assets' sheet starting from row 4."))
        sys.exit(0)

    # ── Validate all rows first ───────────────────────────────────────────────
    print(f"\nValidating rows…")
    valid_rows   = []
    invalid_rows = []

    for raw in rows:
        payload, errors = validate_row(raw)
        if errors:
            invalid_rows.append((raw["_row"], errors))
            print(red(f"  Row {raw['_row']:>4}: ✗  {'; '.join(errors)}"))
        else:
            valid_rows.append((raw["_row"], payload))
            print(green(f"  Row {raw['_row']:>4}: ✓  {payload.get('name', '(unnamed)')}"))

    print(f"\n  {green(str(len(valid_rows)))} valid   "
          f"  {red(str(len(invalid_rows)))} invalid   "
          f"  {len(rows)} total")

    if invalid_rows and not args.skip:
        if not args.dry_run:
            print(red("\nUpload aborted — fix the errors above and re-run."))
            print(dim("  Tip: use --skip to upload valid rows and skip invalid ones."))
        sys.exit(1 if invalid_rows else 0)

    if args.dry_run:
        print(yellow("\nDry run complete — no data was uploaded."))
        sys.exit(0 if not invalid_rows else 1)

    if not valid_rows:
        print(yellow("\nNo valid rows to upload."))
        sys.exit(1)

    # ── Login ─────────────────────────────────────────────────────────────────
    token = login(args.url)

    # ── Upload ────────────────────────────────────────────────────────────────
    print(bold(f"\nUploading {len(valid_rows)} asset(s)…\n"))
    uploaded = 0
    failed   = 0

    for row_num, payload in valid_rows:
        try:
            result = upload_asset(args.url, token, payload)
            print(green(f"  Row {row_num:>4}: ✓  Uploaded → {result.get('asset_id', '?')}  {payload['name']}"))
            uploaded += 1
        except requests.HTTPError as e:
            detail = ""
            try:
                detail = e.response.json().get("detail", str(e))
            except Exception:
                detail = str(e)
            print(red(f"  Row {row_num:>4}: ✗  {payload['name']} — {detail}"))
            failed += 1
        except requests.exceptions.ConnectionError:
            print(red(f"  Row {row_num:>4}: ✗  Connection lost. Check server."))
            failed += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    print(bold(f"\n{'─' * 50}"))
    print(bold(f"  Upload complete"))
    print(f"  {green(str(uploaded))} uploaded successfully")
    if failed:
        print(f"  {red(str(failed))} failed")
    if invalid_rows:
        print(f"  {yellow(str(len(invalid_rows)))} skipped (validation errors)")
    print(f"{'─' * 50}\n")

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
