"""
generate_template.py
Creates the G-Tracker asset bulk-upload Excel template.
Run: python generate_template.py
Output: gtracker_asset_upload_template.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUTPUT = "gtracker_asset_upload_template.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
GOLD_DARK  = "B8960A"
GOLD_LIGHT = "FFF3CD"
HEADER_BG  = "1A1714"
HEADER_FG  = "E8C97A"
INPUT_FG   = "0000FF"
EXAMPLE_FG = "555555"
BORDER_COL = "CCCCCC"

thin  = Side(style="thin",   color=BORDER_COL)
thick = Side(style="medium", color=GOLD_DARK)

def all_border():
    return Border(top=thin, right=thin, bottom=thin, left=thin)

def gold_left_border():
    return Border(top=thin, right=thin, bottom=thin, left=thick)

# ── Column definitions ─────────────────────────────────────────────────────────
# (api_field, header_label, width, note, required)
COLUMNS = [
    ("asset_number",          "ASSET NUMBER",          18, "Leave blank to auto-generate (e.g. AST-00001). Must be unique if supplied.", False),
    ("name",                  "NAME",                  30, "Full descriptive name of the asset.",                                        True),
    ("category",              "CATEGORY",              24, "Must match exactly one of the allowed values — see Reference sheet.",         True),
    ("status",                "STATUS",                20, "Must match exactly one of the allowed values — see Reference sheet.",         True),
    ("location",              "LOCATION",              24, "Physical location, e.g. Building A, Floor 2, Room 201.",                     False),
    ("serial_number",         "SERIAL NUMBER",         18, "Manufacturer serial number or internal asset tag.",                          False),
    ("accountable_department","ACCOUNTABLE DEPARTMENT",26, "Department responsible for this asset.",                                     False),
    ("accountable_person",    "ACCOUNTABLE PERSON",    24, "Full name of the person responsible.",                                       False),
    ("purchase_date",         "PURCHASE DATE",         16, "Format: YYYY-MM-DD (e.g. 2024-01-15).",                                     False),
    ("purchase_value",        "PURCHASE VALUE",        18, "Purchase cost in PHP — numbers only, no ₱ sign or commas.",                  False),
    ("service_life_years",    "SERVICE LIFE (YRS)",    18, "Expected useful life in whole years (e.g. 10).",                             False),
    ("depreciation_method",   "DEPRECIATION METHOD",   24, "Must match exactly one of the allowed values — see Reference sheet.",         False),
    ("depreciation_rate",     "DEPRECIATION RATE (%)", 20, "Annual rate in % — only used when method = custom_rate (e.g. 15 for 15%).", False),
    ("repair_cost",           "REPAIR COST",           18, "Cumulative repair costs in PHP — subtracted from book value.",               False),
    ("notes",                 "NOTES",                 40, "Any additional notes about the asset.",                                      False),
]

# Column letter lookup by api_field
COL_LETTER = {col[0]: get_column_letter(i+1) for i, col in enumerate(COLUMNS)}

CATEGORIES   = ["rooms_facilities","furniture_equipment","vehicles_transport",
                 "it_electronics","maintenance_tools","inventory_consumables"]
STATUSES     = ["available","in_use","maintenance","retired","lost"]
DEPR_METHODS = ["straight_line","declining_balance","custom_rate","none"]

EXAMPLE_ROW = {
    "asset_number":          "",
    "name":                  "Deluxe Room 201",
    "category":              "rooms_facilities",
    "status":                "available",
    "location":              "Building A, Floor 2",
    "serial_number":         "RM-201",
    "accountable_department":"Rooms Division",
    "accountable_person":    "Maria Santos",
    "purchase_date":         "2024-01-15",
    "purchase_value":        "450000",
    "service_life_years":    "20",
    "depreciation_method":   "straight_line",
    "depreciation_rate":     "",
    "repair_cost":           "",
    "notes":                 "Sea view, king bed",
}

NUM_COLS   = len(COLUMNS)
DATA_START = 4
DATA_END   = 503


def make_template():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Assets ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assets"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"   # freeze title + header rows

    last_col_letter = get_column_letter(NUM_COLS)

    # Title row
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A1:{last_col_letter}1")
    t = ws["A1"]
    t.value     = "G-Tracker — Asset Bulk Upload Template"
    t.font      = Font(name="Arial", size=14, bold=True, color=HEADER_FG)
    t.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Header row (row 2)
    ws.row_dimensions[2].height = 30
    for col_idx, (api_field, label, width, note, required) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = f"{label}{' *' if required else ''}"
        cell.font      = Font(name="Arial", size=9, bold=True,
                              color="FFFFFF" if required else "BBBBBB")
        cell.fill      = PatternFill("solid", fgColor="2D2A26" if required else "3D3A36")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = all_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c = Comment(f"{'REQUIRED. ' if required else ''}{note}", "G-Tracker")
        c.width = 260; c.height = 70
        cell.comment = c

    # Example row (row 3)
    ws.row_dimensions[3].height = 22
    for col_idx, (api_field, *_) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=EXAMPLE_ROW.get(api_field, ""))
        cell.font      = Font(name="Arial", size=9, italic=True, color=EXAMPLE_FG)
        cell.fill      = PatternFill("solid", fgColor="F5F5F5")
        cell.alignment = Alignment(vertical="center")
        cell.border    = all_border()
    ws.cell(row=3, column=1).comment = Comment(
        "EXAMPLE ROW — shows expected format.\nReplace with your data from row 4 onwards.",
        "G-Tracker"
    )

    # Data rows 4–503
    for row in range(DATA_START, DATA_END + 1):
        ws.row_dimensions[row].height = 20
        for col_idx, (api_field, _, __, ___, required) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font      = Font(name="Arial", size=10, color=INPUT_FG)
            cell.alignment = Alignment(vertical="center")
            cell.fill      = PatternFill("solid", fgColor="FFFFFF")
            # Gold left border on required columns
            cell.border = gold_left_border() if required else all_border()

    # ── Data validations ──────────────────────────────────────────────────────

    def add_list_dv(col_letter, values, error_msg, error_title):
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(values)}"',
            showDropDown=False,
            error=error_msg,
            errorTitle=error_title,
            showErrorMessage=True,
        )
        dv.sqref = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"
        ws.add_data_validation(dv)

    add_list_dv(COL_LETTER["category"],
                CATEGORIES,
                "Must be one of the allowed categories. See Reference sheet.",
                "Invalid Category")

    add_list_dv(COL_LETTER["status"],
                STATUSES,
                "Must be one of the allowed statuses. See Reference sheet.",
                "Invalid Status")

    add_list_dv(COL_LETTER["depreciation_method"],
                DEPR_METHODS,
                "Must be one of the allowed depreciation methods. See Reference sheet.",
                "Invalid Depreciation Method")

    # Date validation
    dv_date = DataValidation(
        type="date", operator="greaterThan", formula1="DATE(1900,1,1)",
        showDropDown=False,
        error="Enter a valid date in YYYY-MM-DD format (e.g. 2024-01-15).",
        errorTitle="Invalid Date", showErrorMessage=True,
    )
    dv_date.sqref = f"{COL_LETTER['purchase_date']}{DATA_START}:{COL_LETTER['purchase_date']}{DATA_END}"
    ws.add_data_validation(dv_date)

    # Numeric validations
    for field in ("purchase_value", "service_life_years", "depreciation_rate", "repair_cost"):
        dv_num = DataValidation(
            type="decimal", operator="greaterThanOrEqual", formula1="0",
            showDropDown=False,
            error="Must be a non-negative number.",
            errorTitle="Invalid Number", showErrorMessage=True,
        )
        cl = COL_LETTER[field]
        dv_num.sqref = f"{cl}{DATA_START}:{cl}{DATA_END}"
        ws.add_data_validation(dv_num)

    # ── Sheet 2: Reference ────────────────────────────────────────────────────
    ref = wb.create_sheet("Reference")
    ref.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [30, 24, 30, 24, 36]):
        ref.column_dimensions[col].width = w

    def rh(row, col, text):
        c = ref.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = all_border()
        ref.row_dimensions[row].height = 24

    def rv(row, col, text, bg="FFFFFF"):
        c = ref.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", size=10, color="333333")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border    = all_border()
        ref.row_dimensions[row].height = 20

    ref.merge_cells("A1:E1")
    t2 = ref["A1"]
    t2.value     = "G-Tracker — Allowed Values Reference"
    t2.font      = Font(name="Arial", size=13, bold=True, color=HEADER_FG)
    t2.fill      = PatternFill("solid", fgColor=HEADER_BG)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ref.row_dimensions[1].height = 32

    rh(2, 1, "CATEGORY")
    for i, v in enumerate(CATEGORIES, 3):
        rv(i, 1, v, "FFFFFF" if i%2==0 else "F9F9F9")

    rh(2, 2, "STATUS")
    for i, v in enumerate(STATUSES, 3):
        rv(i, 2, v, "FFFFFF" if i%2==0 else "F9F9F9")

    rh(2, 3, "DEPRECIATION METHOD")
    for i, v in enumerate(DEPR_METHODS, 3):
        rv(i, 3, v, "FFFFFF" if i%2==0 else "F9F9F9")

    rh(2, 4, "ASSET NUMBER FORMAT")
    asset_num_notes = [
        "Leave blank → auto-generated",
        "Auto format: AST-00001, AST-00002…",
        "Custom format allowed (e.g. RM-101)",
        "Must be UNIQUE across all assets",
        "Max recommended length: 20 chars",
    ]
    for i, v in enumerate(asset_num_notes, 3):
        rv(i, 4, v, GOLD_LIGHT)

    rh(2, 5, "GENERAL NOTES")
    gen_notes = [
        "* = Required field",
        "Dates: YYYY-MM-DD format",
        "Values: numbers only (no ₱ or commas)",
        "depreciation_rate: % per year (e.g. 15)",
        "Only used when method = custom_rate",
        "Max 500 rows per upload file",
    ]
    for i, v in enumerate(gen_notes, 3):
        rv(i, 5, v, "E8F5E9")

    # ── Sheet 3: Instructions ─────────────────────────────────────────────────
    ins = wb.create_sheet("Instructions")
    ins.sheet_view.showGridLines = False
    ins.column_dimensions["A"].width = 85

    def ir(row, text, bold=False, size=10, color="333333", bg=None, height=18):
        c = ins.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", size=size, bold=bold, color=color)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        ins.row_dimensions[row].height = height

    ir(1,  "G-Tracker — Bulk Asset Upload Instructions",
       bold=True, size=14, color=HEADER_FG, bg=HEADER_BG, height=36)
    ir(2,  "")
    ir(3,  "HOW TO USE THIS TEMPLATE", bold=True, size=11, color=GOLD_DARK, height=22)
    ir(4,  "1.  Fill in the 'Assets' sheet starting from Row 4 (Row 3 is an example — replace or delete it).")
    ir(5,  "2.  Columns with a gold left border (Name, Category, Status) are REQUIRED.")
    ir(6,  "3.  Use the dropdown menus in Category, Status, and Depreciation Method columns.")
    ir(7,  "4.  Asset Number is OPTIONAL — leave it blank to auto-generate AST-00001, AST-00002, etc.")
    ir(8,  "5.  See the 'Reference' sheet for all allowed values and the asset number format rules.")
    ir(9,  "6.  Save as .xlsx format (not .csv or .xls).")
    ir(10, "7.  Run: python upload_assets.py --file your_file.xlsx")
    ir(11, "")
    ir(12, "FIELD DESCRIPTIONS", bold=True, size=11, color=GOLD_DARK, height=22)
    ir(13, "asset_number           — Optional unique ID. Leave blank to auto-generate (AST-NNNNN format). If you supply one, it must not already exist in G-Tracker.")
    ir(14, "name                   — Full descriptive name of the asset (required)")
    ir(15, "category               — Asset category key from the Reference sheet (required)")
    ir(16, "status                 — Current status from the Reference sheet (default: available, required)")
    ir(17, "location               — Physical location, e.g. 'Building A, Floor 2, Room 201'")
    ir(18, "serial_number          — Manufacturer serial number or internal asset tag")
    ir(19, "accountable_department — Department responsible for this asset")
    ir(20, "accountable_person     — Full name of the person responsible")
    ir(21, "purchase_date          — Date purchased in YYYY-MM-DD format, e.g. 2024-01-15")
    ir(22, "purchase_value         — Purchase cost in PHP, numbers only (e.g. 450000, not ₱450,000)")
    ir(23, "service_life_years     — Expected useful life in whole years (e.g. 10)")
    ir(24, "depreciation_method    — straight_line | declining_balance | custom_rate | none")
    ir(25, "depreciation_rate      — Annual depreciation % for custom_rate method only (e.g. 15 for 15%/yr)")
    ir(26, "repair_cost            — Cumulative repair costs in PHP (subtracted from book value)")
    ir(27, "notes                  — Any additional information about the asset")
    ir(28, "")
    ir(29, "UPLOAD SCRIPT USAGE", bold=True, size=11, color=GOLD_DARK, height=22)
    ir(30, "python upload_assets.py --file gtracker_asset_upload_template.xlsx [OPTIONS]")
    ir(31, "")
    ir(32, "Options:")
    ir(33, "  --url     Base URL of G-Tracker (default: http://localhost:8000)")
    ir(34, "  --file    Path to the filled-in Excel file")
    ir(35, "  --sheet   Sheet name to read from (default: Assets)")
    ir(36, "  --dry-run Validate rows without uploading anything")
    ir(37, "  --skip    Skip invalid rows and continue uploading valid ones")
    ir(38, "")
    ir(39, "Examples:")
    ir(40, "  python upload_assets.py --file my_assets.xlsx --dry-run")
    ir(41, "  python upload_assets.py --file my_assets.xlsx --url https://g.endonalaw.com")
    ir(42, "  python upload_assets.py --file my_assets.xlsx --url https://g.endonalaw.com --skip")
    ir(43, "")
    ir(44, "The script prompts for your G-Tracker username and password. Credentials are never saved to disk.")

    wb.save(OUTPUT)
    print(f"Template saved: {OUTPUT}")


if __name__ == "__main__":
    make_template()
